from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework_simplejwt.tokens import RefreshToken
from django.contrib.auth.models import User
from django.contrib.auth import authenticate
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework_simplejwt.views import TokenObtainPairView, TokenRefreshView
from rest_framework_simplejwt.authentication import JWTAuthentication
from .models import *
from .serializers import *
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from .utils import generate_equipment_pdf
from django.core.mail import EmailMessage
from django.conf import settings
import binascii
from openpyxl import load_workbook

# Custom Token View with better error handling
class CustomTokenObtainPairView(TokenObtainPairView):
    permission_classes = [AllowAny]

    def post(self, request, *args, **kwargs):
        try:
            response = super().post(request, *args, **kwargs)
            return response
        except Exception as e:
            return Response(
                {"error": "Invalid credentials"},
                status=status.HTTP_401_UNAUTHORIZED
            )

# Logout View with token blacklisting
class LogoutView(APIView):
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication]

    def post(self, request):
        try:
            refresh_token = request.data.get('refresh_token')
            if refresh_token:
                token = RefreshToken(refresh_token)
                token.blacklist()
            return Response({"message": "Successfully logged out"}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": "Invalid token"}, status=status.HTTP_400_BAD_REQUEST)

# Protected DataCenter List View
class DataCenterListView(APIView):
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication]

    def get(self, request):
        try:
            data_centers = DataCenter.objects.all()
            serializer = DataCenterSerializer(data_centers, many=True)
            return Response(serializer.data)
        except Exception as e:
            return Response(
                {"error": "Failed to fetch datacenters"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

# Protected DataCenter Detail View
class DataCenterDetailView(APIView):
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication]

    def get(self, request, pk):
        try:
            data_center = DataCenter.objects.get(pk=pk)
            serializer = DataCenterSerializer(data_center)
            return Response(serializer.data)
        except DataCenter.DoesNotExist:
            return Response(
                {"error": "DataCenter not found"},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {"error": "Failed to fetch datacenter details"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
class EquipmentFetchView(APIView):
    def get(self, request, datacenter_id):
        # Get query parameters for search
        service_tag = request.GET.get('service_tag', '').strip()
        license_type = request.GET.get('license_type', '').strip()

        try:
            # Get the DataCenter by ID
            datacenter = DataCenter.objects.get(pk=datacenter_id)

            # Start with all equipment for this datacenter
            equipments = Equipment.objects.filter(datacenter=datacenter)

            # Apply search filters if provided
            if service_tag:
                equipments = equipments.filter(service_tag__icontains=service_tag)
            if license_type:
                equipments = equipments.filter(license_type__icontains=license_type)

            # Serialize the equipment data
            serializer = EquipmentSerializer(equipments, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)

        except DataCenter.DoesNotExist:
            return Response({"error": "DataCenter not found"}, status=status.HTTP_404_NOT_FOUND)

class EquipmentAddToDataCenterView(APIView):
    def post(self, request, datacenter_id):
        try:
            # Get the DataCenter by ID
            datacenter = DataCenter.objects.get(pk=datacenter_id)

            # Add the datacenter to the serializer context
            serializer = AddEquipmentSerializer(data=request.data, context={'datacenter': datacenter})
            
            if serializer.is_valid():
                # Save the equipment instance
                equipment = serializer.save()
                return Response(AddEquipmentSerializer(equipment).data, status=status.HTTP_201_CREATED)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        except DataCenter.DoesNotExist:
            return Response({"error": "DataCenter not found"}, status=status.HTTP_404_NOT_FOUND)
        

class EquipmentModifyView(APIView):
    def patch(self, request, datacenter_id, equipment_id):
        try:
            # Check if the DataCenter exists
            datacenter = DataCenter.objects.get(pk=datacenter_id)

            # Retrieve the Equipment instance by its ID
            equipment = Equipment.objects.get(pk=equipment_id, datacenter=datacenter)

            # Use the ModifyEquipmentSerializer to validate and update the data
            serializer = ModifyEquipmentSerializer(equipment, data=request.data, partial=True)
            
            if serializer.is_valid():
                # Save the updated equipment instance
                serializer.save()
                return Response(serializer.data, status=status.HTTP_200_OK)

            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        except DataCenter.DoesNotExist:
            return Response({"error": "DataCenter not found"}, status=status.HTTP_404_NOT_FOUND)
        except Equipment.DoesNotExist:
            return Response({"error": "Equipment not found in this DataCenter"}, status=status.HTTP_404_NOT_FOUND)
        

class EquipmentDeleteView(APIView):
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication]
    
    def delete(self, request, datacenter_id, equipment_id):
        try:
            # Check if the DataCenter exists
            datacenter = DataCenter.objects.get(pk=datacenter_id)

            # Retrieve the Equipment instance by its ID and check if it belongs to the specified DataCenter
            equipment = Equipment.objects.get(pk=equipment_id, datacenter=datacenter)
            
            # Store equipment info for response
            equipment_info = {
                'id': equipment.id,
                'equipment_type': equipment.equipment_type,
                'service_tag': equipment.service_tag
            }
            
            # Delete the equipment
            equipment.delete()

            # Log the deletion
            print(f"Equipment {equipment_info['service_tag']} deleted by {request.user.username}")
            
            # Return success response with 200 OK and message
            return Response({
                "success": True,
                "message": "Equipment deleted successfully.",
                "deleted_equipment": equipment_info
            }, status=status.HTTP_200_OK)

        except DataCenter.DoesNotExist:
            return Response({
                "success": False,
                "error": "DataCenter not found"
            }, status=status.HTTP_404_NOT_FOUND)
            
        except Equipment.DoesNotExist:
            return Response({
                "success": False,
                "error": "Equipment not found in this DataCenter"
            }, status=status.HTTP_404_NOT_FOUND)
            
        except Exception as e:
            return Response({
                "success": False,
                "error": f"An error occurred: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class EquipmentLicenseTypeAutocompleteView(APIView):
    def get(self, request, datacenter_id):
        try:
            # Get the DataCenter by ID
            datacenter = DataCenter.objects.get(pk=datacenter_id)

            # Get all distinct license types for the equipments in the given DataCenter
            license_types = Equipment.objects.filter(datacenter=datacenter).values_list('license_type', flat=True).distinct()

            # Return the license types as a list
            return Response(list(license_types), status=status.HTTP_200_OK)

        except DataCenter.DoesNotExist:
            return Response({"error": "DataCenter not found"}, status=status.HTTP_404_NOT_FOUND)
        
class EquipmentServiceTagAutocompleteView(APIView):
    def get(self, request, datacenter_id):
        try:
            # Get the DataCenter by ID
            datacenter = DataCenter.objects.get(pk=datacenter_id)

            # Get all distinct service tags for the equipments in the given DataCenter
            service_tags = Equipment.objects.filter(datacenter=datacenter).values_list('service_tag', flat=True).distinct()

            # Return the service tags as a list
            return Response(list(service_tags), status=status.HTTP_200_OK)

        except DataCenter.DoesNotExist:
            return Response({"error": "DataCenter not found"}, status=status.HTTP_404_NOT_FOUND)
        
class EquipmentExportExcelView(APIView):
    def get(self, request, datacenter_id):
        # Get query parameters for filtering
        service_tag = request.GET.get('service_tag', None)
        license_type = request.GET.get('license_type', None)

        try:
            # Get the DataCenter by ID
            datacenter = DataCenter.objects.get(pk=datacenter_id)

            # Get all equipments for the given DataCenter
            equipments = Equipment.objects.filter(datacenter=datacenter)

            # Apply filters if provided
            if service_tag:
                equipments = equipments.filter(service_tag__icontains=service_tag)
            if license_type:
                equipments = equipments.filter(license_type__icontains=license_type)

            # Create a new Excel workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Equipments"

            # Add headers to the worksheet
            headers = ["ID", "Equipment Type", "Service Tag", "License Type", "Serial Number", "License Expiry Date"]
            for col_num, header in enumerate(headers, 1):
                col_letter = get_column_letter(col_num)
                ws[f"{col_letter}1"] = header

            # Add equipment data to the worksheet
            for row_num, equipment in enumerate(equipments, 2):
                ws[f"A{row_num}"] = equipment.id
                ws[f"B{row_num}"] = equipment.equipment_type
                ws[f"C{row_num}"] = equipment.service_tag
                ws[f"D{row_num}"] = equipment.license_type
                ws[f"E{row_num}"] = equipment.serial_number
                ws[f"F{row_num}"] = equipment.license_expired_date

            # Create the HttpResponse for file download
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response['Content-Disposition'] = f'attachment; filename="equipments_{datacenter_id}.xlsx"'

            # Save the workbook to the HttpResponse
            wb.save(response)
            return response

        except DataCenter.DoesNotExist:
            return Response({"error": "DataCenter not found"}, status=404)
        

class EquipmentExportPDFView(APIView):
    def get(self, request, datacenter_id):
        # Get query parameters for filtering
        service_tag = request.GET.get('service_tag', None)
        license_type = request.GET.get('license_type', None)

        try:
            # Get the DataCenter by ID
            datacenter = DataCenter.objects.get(pk=datacenter_id)

            # Get all equipments for the given DataCenter
            equipments = Equipment.objects.filter(datacenter=datacenter)

            # Apply filters if provided
            if service_tag:
                equipments = equipments.filter(service_tag__icontains=service_tag)
            if license_type:
                equipments = equipments.filter(license_type__icontains=license_type)

            # Generate the PDF file
            pdf_buffer = generate_equipment_pdf(equipments)

            # Create the HTTP Response for file download
            response = HttpResponse(pdf_buffer, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="equipments_{datacenter_id}.pdf"'
            return response

        except DataCenter.DoesNotExist:
            return Response({"error": "DataCenter not found"}, status=404)
        except Exception as e:
            return Response({"error": str(e)}, status=500)

class EquipmentImportExcelView(APIView):
    def post(self, request, datacenter_id):
        print("\n=== Excel Import Request Received ===")
        print(f"Datacenter ID: {datacenter_id}")
        print(f"Request user: {request.user}" if hasattr(request, 'user') else "No user in request")
        print(f"Request FILES: {request.FILES}")
        
        # Check if user is authenticated
        if not hasattr(request, 'user') or not request.user.is_authenticated:
            print("User not authenticated")
            return Response({"error": "Authentication required"}, status=401)
            
        # Check if file is present in the request
        if 'file' not in request.FILES:
            error_msg = "No file provided in the request"
            print(error_msg)
            return Response({"error": error_msg}, status=400)
            
        excel_file = request.FILES['file']
        print(f"Processing file: {excel_file.name} ({excel_file.size} bytes)")
        
        try:
            # Check file extension
            if not excel_file.name.lower().endswith(('.xlsx', '.xls')):
                error_msg = f"Invalid file type: {excel_file.name}. Please upload a valid Excel file (.xlsx or .xls)."
                print(error_msg)
                return Response({"error": error_msg}, status=400)

            try:
                # Read the Excel file
                wb = load_workbook(excel_file, data_only=True)
                ws = wb.active

                # Get headers from first row and create a mapping of lowercase header to column index
                headers = []
                header_map = {}
                for idx, cell in enumerate(ws[1]):
                    header = str(cell.value).strip().lower() if cell.value else f"column_{idx+1}"
                    headers.append(header)
                    header_map[header] = idx

                # Map headers to their positions
                header_positions = {}
                for idx, header in enumerate(headers):
                    header_lower = str(header).lower().strip()
                    header_positions[header_lower] = idx

                # Map of possible column names to their standard names and positions
                column_mapping = {
                    'equipment type': ('equipment_type', 'equipment type', 'type', 'equipment'),
                    'service tag': ('service_tag', 'service tag', 'service', 'tag'),
                    'license type': ('license_type', 'license type', 'license'),
                    'serial number': ('serial_number', 'serial number', 'serial', 'sn'),
                    'license expiry date': ('license_expired_date', 'license expiry', 'expiry date', 'expires', 'expiry')
                }

                # Find column positions
                column_positions = {}
                for display_name, aliases in column_mapping.items():
                    for alias in aliases:
                        if alias in header_positions:
                            column_positions[display_name] = header_positions[alias]
                            break

                # Check for required columns
                required_columns = ['equipment type', 'service tag', 'license type', 'serial number']
                missing_columns = [col for col in required_columns if col not in column_positions]
                
                if missing_columns:
                    return Response({
                        "error": "Missing required columns in the Excel file.",
                        "missing_columns": missing_columns,
                        "available_columns": headers
                    }, status=400)

                # Map of display names to model fields
                field_mapping = {
                    'equipment type': {'field': 'equipment_type', 'required': True},
                    'service tag': {'field': 'service_tag', 'required': True},
                    'license type': {'field': 'license_type', 'required': True},
                    'serial number': {'field': 'serial_number', 'required': True},
                    'license expiry date': {'field': 'license_expired_date', 'required': False}
                }


                new_equipments = []
                updated_count = 0
                error_messages = []
                
                # Get the datacenter object
                try:
                    datacenter = DataCenter.objects.get(pk=datacenter_id)
                    print(f"Found datacenter: {datacenter.name} (ID: {datacenter.id})")
                except DataCenter.DoesNotExist:
                    return Response({"error": f"DataCenter with ID {datacenter_id} not found"}, status=404)

                # Process each row starting from row 2
                print("\n=== Starting to process rows ===")
                for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
                    try:
                        print(f"\n--- Processing row {row_num} ---")
                        row_values = [cell.value for cell in row]
                        print(f"Row values: {row_values}")
                        
                        # Skip empty rows
                        if all(cell.value is None for cell in row):
                            print(f"Row {row_num}: Skipping empty row")
                            continue
                            
                        equipment_data = {}
                        
                        # Extract data based on column positions
                        print("Extracting data from columns:")
                        for display_name, field_info in field_mapping.items():
                            field_name = field_info['field']
                            is_required = field_info['required']
                            
                            if display_name in column_positions:
                                col_idx = column_positions[display_name]
                                if col_idx < len(row):
                                    value = row[col_idx].value
                                    print(f"  {display_name} (col {col_idx}): {value} (type: {type(value).__name__ if value else 'None'})")
                                    
                                    # Skip None values for non-required fields
                                    if value is None and not is_required:
                                        print(f"  - Skipping optional field {field_name}: None")
                                        continue
                                        
                                    if value is not None:
                                        # Convert to string for text fields
                                        if field_name != 'license_expired_date':
                                            equipment_data[field_name] = str(value).strip()
                                            print(f"  - Set {field_name} = {equipment_data[field_name]}")
                                        # Handle date field
                                        elif value:
                                            if hasattr(value, 'strftime'):  # Already a date object
                                                equipment_data[field_name] = value.date() if hasattr(value, 'date') else value
                                                print(f"  - Set date from date object: {equipment_data[field_name]}")
                                            else:
                                                try:
                                                    # Try to parse date string
                                                    from datetime import datetime
                                                    value_str = str(value).strip()
                                                    if value_str:  # Only try to parse if not empty string
                                                        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%Y/%m/%d', '%d-%m-%Y', '%m-%d-%Y'):
                                                            try:
                                                                equipment_data[field_name] = datetime.strptime(value_str, fmt).date()
                                                                print(f"  - Parsed date '{value_str}' as {equipment_data[field_name]} with format '{fmt}'")
                                                                break
                                                            except ValueError:
                                                                continue
                                                except Exception as e:
                                                    print(f"Error parsing date '{value}': {str(e)}")
                                                    equipment_data[field_name] = None
                            else:
                                print(f"  {display_name}: Column not found")
                        
                        # Skip empty rows
                        print(f"Equipment data extracted: {equipment_data}")
                        if not any(equipment_data.values()):
                            error_msg = f"Row {row_num}: Empty row skipped"
                            error_messages.append(error_msg)
                            print(error_msg)
                            continue
                            
                        print(f"Row {row_num} data after extraction:")
                        for k, v in equipment_data.items():
                            print(f"  {k}: {v} (type: {type(v).__name__ if v else 'None'})")

                        # Debug log the row data with column names
                        print(f"\nProcessing row {row_num}:")
                        for col_name, value in equipment_data.items():
                            print(f"  {col_name}: {value} (type: {type(value).__name__})")

                        # Validate required fields with more details
                        required_fields = {
                            'equipment_type': 'Equipment Type',
                            'service_tag': 'Service Tag',
                            'license_type': 'License Type',
                            'serial_number': 'Serial Number'
                        }
                        
                        # Make license_expired_date optional and set a default future date if not provided
                        if 'license_expired_date' not in equipment_data or equipment_data['license_expired_date'] is None:
                            from datetime import datetime, timedelta
                            equipment_data['license_expired_date'] = (datetime.now() + timedelta(days=365)).date()  # Default to 1 year from now
                        
                        missing_fields = []
                        for field, display_name in required_fields.items():
                            value = equipment_data.get(field)
                            if not value and value != 0:  # 0 is valid for some fields
                                missing_fields.append(display_name)
                            elif isinstance(value, str) and not value.strip():
                                missing_fields.append(f"{display_name} (empty string)")
                        
                        if missing_fields:
                            error_msg = (f"Row {row_num}: Missing or empty required fields: {', '.join(missing_fields)}. "
                                        f"Available columns: {', '.join(equipment_data.keys())}")
                            error_messages.append(error_msg)
                            print(f"ERROR: {error_msg}")
                            continue

                        # Just validate the serial number is present
                        try:
                            serial = str(equipment_data.get('serial_number', '')).strip()
                            if not serial:
                                error_msg = f"Row {row_num}: Empty serial number"
                                error_messages.append(error_msg)
                                print(error_msg)
                                continue
                        except Exception as e:
                            error_msg = f"Row {row_num}: Invalid serial number - {str(e)}"
                            error_messages.append(error_msg)
                            print(error_msg)
                            continue

                        try:
                            serial = str(equipment_data['serial_number']).strip()
                            
                            # Check if equipment with this serial number already exists
                            existing_equipment = Equipment.objects.filter(serial_number=serial).first()
                            
                            if existing_equipment:
                                # Update existing equipment
                                existing_equipment.equipment_type = str(equipment_data['equipment_type']).strip()
                                existing_equipment.service_tag = str(equipment_data['service_tag']).strip()
                                existing_equipment.license_type = str(equipment_data['license_type']).strip()
                                existing_equipment.license_expired_date = equipment_data['license_expired_date']
                                existing_equipment.datacenter = datacenter
                                existing_equipment.full_clean()
                                existing_equipment.save()
                                print(f"Row {row_num}: Updated existing equipment - {existing_equipment}")
                                updated_count += 1
                            else:
                                # Create new equipment
                                equipment = Equipment(
                                    equipment_type=str(equipment_data['equipment_type']).strip(),
                                    service_tag=str(equipment_data['service_tag']).strip(),
                                    license_type=str(equipment_data['license_type']).strip(),
                                    serial_number=serial,
                                    license_expired_date=equipment_data['license_expired_date'],
                                    datacenter=datacenter
                                )
                                equipment.full_clean()
                                equipment.save()
                                new_equipments.append(equipment)
                                print(f"Row {row_num}: Added new equipment - {equipment}")
                                
                            print(f"Assigned to datacenter: {datacenter.name} (ID: {datacenter.id})")
                        except Exception as e:
                            error_msg = f"Row {row_num}: Error creating equipment - {str(e)}"
                            error_messages.append(error_msg)
                            print(error_msg)

                    except Exception as e:
                        error_messages.append(f"Row {row_num}: {str(e)}")
                        continue

                # Create each equipment individually to handle duplicates
                created_count = 0
                for equipment in new_equipments:
                    try:
                        equipment.save()
                        created_count += 1
                    except Exception as e:
                        error_messages.append(f"Error saving equipment {equipment.service_tag}: {str(e)}")
                        continue

                # Prepare response
                response_data = {
                    "message": f"Successfully processed {len(new_equipments) + updated_count} equipment items ({len(new_equipments)} new, {updated_count} updated).",
                    "imported_count": len(new_equipments),
                    "updated_count": updated_count,
                    "error_count": len(error_messages),
                }
                
                if error_messages:
                    response_data["errors"] = error_messages
                    response_data["suggestions"] = [
                        "Please check that all required fields are present and correctly formatted.",
                        "Ensure serial numbers are unique.",
                        "Check that date fields are in a valid format (YYYY-MM-DD, DD/MM/YYYY, or MM/DD/YYYY)."
                    ]

                return Response(response_data, status=status.HTTP_200_OK)

            except Exception as e:
                import traceback
                error_trace = traceback.format_exc()
                print(f"Error processing Excel file: {error_trace}")
                return Response({
                    "error": "Error processing Excel file",
                    "details": str(e),
                    "suggestion": "Please check the file format and ensure all required columns are present.",
                    "required_columns": ["Equipment Type", "Service Tag", "License Type", "Serial Number"],
                    "trace": error_trace if settings.DEBUG else None
                }, status=400)

        except DataCenter.DoesNotExist:
            return Response({"error": "DataCenter not found"}, status=404)
        except Exception as e:
            import traceback
            error_trace = traceback.format_exc()
            print(f"Unexpected error: {error_trace}")
            return Response({
                "error": "An unexpected error occurred",
                "details": str(e),
                "trace": error_trace if settings.DEBUG else None
            }, status=500)


class EquipmentSendPDFByEmailView(APIView):
    def post(self, request, datacenter_id):
        import binascii
        # Use email from frontend if provided, else fallback
        recipient = request.data.get('email', 'test@example.com')
        try:
            datacenter = DataCenter.objects.get(pk=datacenter_id)
            equipments = Equipment.objects.filter(datacenter=datacenter)
            if not equipments.exists():
                return Response({'error': 'No equipment found for this datacenter.'}, status=404)
            pdf_buffer = generate_equipment_pdf(equipments)
            pdf_bytes = pdf_buffer.getvalue()
            print(f'[DEBUG] Generated PDF size: {len(pdf_bytes)} bytes')
            print(f'[DEBUG] PDF header: {binascii.hexlify(pdf_bytes[:16])}')
            if not pdf_bytes or len(pdf_bytes) < 100:
                print('[DEBUG] PDF is empty or too small!')
                raise Exception('Generated PDF is empty or too small!')
            msg = EmailMessage(
                subject=f"Equipment Information for {datacenter.name}",
                body=f"Please find attached the equipment information for datacenter: {datacenter.name}.",
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=[recipient],
            )
            msg.attach(f"equipments_{datacenter.name}.pdf", pdf_bytes, 'application/pdf')
            msg.send(fail_silently=False)
            print(f'[DEBUG] Sent real PDF to {recipient}')
            return Response({'message': f'PDF sent successfully to {recipient}!'}, status=200)
        except DataCenter.DoesNotExist:
            return Response({'error': 'Datacenter not found.'}, status=404)
        except Exception as e:
            import traceback
            print('[DEBUG] Exception in PDF email sending:')
            traceback.print_exc()
            return Response({'error': str(e)}, status=500)