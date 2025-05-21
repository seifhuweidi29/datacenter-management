from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Create a new workbook
wb = Workbook()
ws = wb.active

# Define headers
headers = [
    'equipment_type',
    'service_tag',
    'license_type',
    'serial_number',
    'license_expired_date'
]

# Add headers with formatting
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Add example row
example_row = 2
ws.cell(row=example_row, column=1, value="Server")
ws.cell(row=example_row, column=2, value="ST001")
ws.cell(row=example_row, column=3, value="Standard")
ws.cell(row=example_row, column=4, value="SN001")
ws.cell(row=example_row, column=5, value="2025-05-20")

# Add instructions
instructions = [
    "# Instructions:",
    "# 1. Keep all headers in lowercase and in this exact order",
    "# 2. Add your equipment data below this line",
    "# 3. Each row must have all 5 fields filled",
    "# 4. Serial numbers must be unique across all equipment",
    "# 5. Dates must be in YYYY-MM-DD format",
    "# 6. Service tags should follow the pattern: STXXX",
    "# 7. Serial numbers should follow the pattern: SNXXX",
    "",
    "# Valid Equipment Types:",
    "# - Server",
    "# - Switch",
    "# - Router",
    "# - Firewall",
    "# - Storage",
    "# - Load Balancer",
    "",
    "# Valid License Types:",
    "# - Standard",
    "# - Advanced",
    "# - Enterprise",
    "# - Premium"
]

# Add instructions starting from row 4
for i, line in enumerate(instructions, 4):
    ws.cell(row=i, column=1, value=line)
    ws.cell(row=i, column=1).font = Font(color="FF0000")  # Red color for instructions

# Auto-adjust column widths
for col in range(1, len(headers) + 1):
    max_length = 0
    column = get_column_letter(col)
    for cell in ws[column]:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

# Save the file
wb.save('/home/seif/datacenter_project/datacenter_app/equipment_import_template.xlsx')
