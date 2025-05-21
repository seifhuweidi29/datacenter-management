import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from django.core.management.base import BaseCommand

class Command(BaseCommand):
    help = 'Generate sample Excel file for equipment import'

    def handle(self, *args, **options):
        # Create a new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Define headers
        headers = [
            'Equipment Type',
            'Service Tag',
            'License Type',
            'Serial Number',
            'License Expiry Date'
        ]
        
        # Add headers
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)

        # Generate sample data
        equipment_types = ['Server', 'Switch', 'Router', 'Firewall', 'Storage', 'Load Balancer']
        license_types = ['Standard', 'Advanced', 'Enterprise', 'Premium']
        
        for row in range(2, 25):  # Create 23 rows (20+ equipments)
            # Generate sample data
            equipment_type = equipment_types[(row-2) % len(equipment_types)]
            service_tag = f'ST{row:03d}'
            license_type = license_types[(row-2) % len(license_types)]
            serial_number = f'SN{row:03d}'
            # Generate random expiry date between now and 2 years from now
            days_offset = (row-2) * 30  # Different expiry dates for each row
            expiry_date = (datetime.now() + timedelta(days=days_offset)).date()
            
            # Add data to row
            ws.cell(row=row, column=1, value=equipment_type)
            ws.cell(row=row, column=2, value=service_tag)
            ws.cell(row=row, column=3, value=license_type)
            ws.cell(row=row, column=4, value=serial_number)
            ws.cell(row=row, column=5, value=expiry_date.strftime('%Y-%m-%d'))

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            max_length = 0
            column = get_column_letter(col)
            for cell in ws[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Save the file
        file_path = '/home/seif/datacenter_project/datacenter_app/datacenter_equipment.xlsx'
        wb.save(file_path)
        self.stdout.write(self.style.SUCCESS(f'Successfully created sample Excel file at {file_path}'))
